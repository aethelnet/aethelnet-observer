"""
Distill - Universal Document Processor

Extract signal from noise in any document format:
- AI chat logs (Claude, Gemini, GPT) → structured decisions + actions
- Emails/Slack → tasks, commitments, key info
- PDFs/docs → headings, tables, key points
- Code reviews → actionable feedback

Usage:
    from services.distill import Distiller
    
    distiller = Distiller()
    blocks = await distiller.process_file('/path/to/chat.txt')
    blocks = await distiller.process_content(text, format='ai_chat')
"""

import re
import os
from typing import List, Dict, Optional, Any
from dataclasses import dataclass, asdict, field
from enum import Enum


class SignalType(Enum):
    """Classification of extracted signal."""
    ACTION = "action"           # Something to do
    DECISION = "decision"       # Choice that was made
    INFO = "info"               # Key information
    CODE = "code"               # Code snippet
    ERROR = "error"             # Problem/bug
    QUESTION = "question"       # Unanswered question
    CONTEXT = "context"         # Background info
    NOISE = "noise"             # Repetition/alignment
    # Data analysis types
    DATA = "data"               # Raw data point/row
    METRIC = "metric"           # Calculated metric (mean, sum, etc.)
    TREND = "trend"             # Detected trend/pattern
    OUTLIER = "outlier"         # Anomaly/outlier detection
    CORRELATION = "correlation" # Relationship between variables
    

@dataclass
class DistilledBlock:
    """A distilled piece of signal from a document."""
    content: str
    signal_type: SignalType
    confidence: float           # 0.0 - 1.0
    source_line: int
    source_format: str          # ai_chat, email, markdown, code
    tags: List[str] = field(default_factory=list)
    metadata: Dict[str, Any] = field(default_factory=dict)
    
    def to_dict(self) -> dict:
        d = asdict(self)
        d['signal_type'] = self.signal_type.value
        return d


class Distiller:
    """
    Universal document processor that extracts signal from noise.
    """
    
    # AI Chat patterns
    AI_CHAT_PATTERNS = {
        'action': [
            r'(?:TODO|FIXME|ACTION|NEXT STEP)[:\s]+(.+)',
            r'(?:I will|Let me|I\'ll|We should|We need to)\s+(.+)',
            r'(?:Creating|Implementing|Adding|Fixing|Updating)\s+(.+)',
        ],
        'decision': [
            r'(?:Decision|Decided|Choosing|We\'ll go with)[:\s]+(.+)',
            r'(?:The plan is|Our approach|Strategy)[:\s]+(.+)',
        ],
        'code': [
            r'```[\w]*\n([\s\S]+?)```',
        ],
        'error': [
            r'(?:Error|Bug|Issue|Problem|Failed)[:\s]+(.+)',
            r'(?:doesn\'t work|not working|broken|failing)',
        ],
        'question': [
            r'\?(?:\s|$)',
            r'(?:Should we|Could we|How do we|What if)',
        ],
    }
    
    # Email patterns
    EMAIL_PATTERNS = {
        'action': [
            r'(?:Please|Kindly|Could you)\s+(.+)',
            r'(?:Deadline|Due by|By EOD)[:\s]+(.+)',
        ],
        'info': [
            r'(?:FYI|For your information|Note that)[:\s]+(.+)',
        ],
    }
    
    # Noise patterns (to filter out)
    NOISE_PATTERNS = [
        r'^(Ok|Okay|Sure|Got it|Thanks|Thank you|I understand)\.?$',
        r'^(Yes|No|Agreed|Correct)\.?$',
        r'^I\'ll (help|assist) you with that\.?$',
        r'^(Hello|Hi|Hey)!?\s*$',
        r'^(Let me|I\'ll) (continue|proceed|start)\.?$',
        r'^Of course\.?$',
    ]
    
    def __init__(self):
        self.noise_re = [re.compile(p, re.IGNORECASE) for p in self.NOISE_PATTERNS]
    
    async def process_file(self, file_path: str) -> List[DistilledBlock]:
        """Process any file and extract signal."""
        if not os.path.isfile(file_path):
            return []
        
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
            content = f.read()
        
        # Detect format from extension or content
        format_type = self._detect_format(file_path, content)
        return await self.process_content(content, format_type)
    
    async def process_content(
        self, 
        content: str, 
        format_type: str = 'auto'
    ) -> List[DistilledBlock]:
        """Process content string and extract signal."""
        if format_type == 'auto':
            format_type = self._detect_format_from_content(content)
        
        if format_type == 'ai_chat':
            return self._process_ai_chat(content)
        elif format_type == 'email':
            return self._process_email(content)
        elif format_type == 'markdown':
            return self._process_markdown(content)
        elif format_type == 'code':
            return self._process_code(content)
        elif format_type == 'csv':
            return self._process_csv(content)
        elif format_type == 'json':
            return self._process_json(content)
        else:
            return self._process_generic(content)
    
    def _detect_format(self, file_path: str, content: str) -> str:
        """Detect document format from path and content."""
        ext = os.path.splitext(file_path)[1].lower()
        
        # Data analysis formats (deterministic, no LLM needed)
        if ext in ['.xlsx', '.xls', '.ods']:
            return 'spreadsheet'
        elif ext in ['.csv', '.tsv']:
            return 'csv'
        elif ext == '.json':
            return 'json'
        elif ext in ['.parquet', '.feather']:
            return 'dataframe'
        
        # Code formats
        if ext in ['.py', '.js', '.ts', '.vue', '.java', '.cpp']:
            return 'code'
        elif ext == '.md':
            return 'markdown'
        elif ext in ['.eml', '.msg']:
            return 'email'
        
        # Check content patterns
        return self._detect_format_from_content(content)
    
    def _detect_format_from_content(self, content: str) -> str:
        """Detect format from content patterns."""
        # AI chat indicators
        ai_indicators = [
            'Claude:', 'Gemini:', 'GPT:', 'Assistant:', 'AI:',
            'Human:', 'User:', '<thinking>', '</thinking>',
            'Step Id:', '<function_calls>'
        ]
        if any(ind in content for ind in ai_indicators):
            return 'ai_chat'
        
        # Email indicators
        if 'From:' in content and 'Subject:' in content:
            return 'email'
        
        # Markdown indicators
        if content.strip().startswith('#') or '\n## ' in content:
            return 'markdown'
        
        return 'generic'
    
    def _is_noise(self, text: str) -> bool:
        """Check if text is alignment noise."""
        text = text.strip()
        for pattern in self.noise_re:
            if pattern.match(text):
                return True
        return False
    
    def _process_ai_chat(self, content: str) -> List[DistilledBlock]:
        """Process AI chat logs."""
        blocks = []
        lines = content.split('\n')
        
        current_speaker = None
        current_block = []
        block_start = 0
        
        for i, line in enumerate(lines, 1):
            # Detect speaker changes
            speaker_match = re.match(r'^(Human|User|Assistant|Claude|Gemini|GPT|AI):', line)
            if speaker_match:
                # Flush previous block
                if current_block:
                    blocks.extend(self._analyze_chat_block(
                        '\n'.join(current_block), 
                        current_speaker,
                        block_start
                    ))
                current_speaker = speaker_match.group(1)
                current_block = [line[len(speaker_match.group(0)):].strip()]
                block_start = i
            else:
                current_block.append(line)
        
        # Flush last block
        if current_block:
            blocks.extend(self._analyze_chat_block(
                '\n'.join(current_block),
                current_speaker,
                block_start
            ))
        
        return blocks
    
    def _analyze_chat_block(
        self, 
        text: str, 
        speaker: Optional[str],
        line_num: int
    ) -> List[DistilledBlock]:
        """Analyze a chat block for signal."""
        blocks = []
        text = text.strip()
        
        if not text:
            return blocks
        
        # Check for noise
        if self._is_noise(text):
            return [DistilledBlock(
                content=text,
                signal_type=SignalType.NOISE,
                confidence=0.9,
                source_line=line_num,
                source_format='ai_chat',
                tags=['skip']
            )]
        
        # Extract code blocks
        code_matches = re.findall(r'```(\w*)\n([\s\S]+?)```', text)
        for lang, code in code_matches:
            blocks.append(DistilledBlock(
                content=code.strip(),
                signal_type=SignalType.CODE,
                confidence=1.0,
                source_line=line_num,
                source_format='ai_chat',
                tags=[lang] if lang else ['code'],
                metadata={'language': lang}
            ))
        
        # Remove code blocks for further analysis
        text_no_code = re.sub(r'```[\s\S]+?```', '', text)
        
        # Check for actions
        for pattern in self.AI_CHAT_PATTERNS['action']:
            matches = re.findall(pattern, text_no_code, re.IGNORECASE)
            for match in matches:
                blocks.append(DistilledBlock(
                    content=match.strip(),
                    signal_type=SignalType.ACTION,
                    confidence=0.8,
                    source_line=line_num,
                    source_format='ai_chat',
                    tags=['action']
                ))
        
        # Check for decisions
        for pattern in self.AI_CHAT_PATTERNS['decision']:
            matches = re.findall(pattern, text_no_code, re.IGNORECASE)
            for match in matches:
                blocks.append(DistilledBlock(
                    content=match.strip(),
                    signal_type=SignalType.DECISION,
                    confidence=0.7,
                    source_line=line_num,
                    source_format='ai_chat',
                    tags=['decision']
                ))
        
        # Check for errors
        for pattern in self.AI_CHAT_PATTERNS['error']:
            if re.search(pattern, text_no_code, re.IGNORECASE):
                blocks.append(DistilledBlock(
                    content=text_no_code[:200],
                    signal_type=SignalType.ERROR,
                    confidence=0.6,
                    source_line=line_num,
                    source_format='ai_chat',
                    tags=['error', 'needs_attention']
                ))
                break
        
        # If no specific signal found, classify as info
        if not blocks and len(text_no_code) > 50:
            blocks.append(DistilledBlock(
                content=text_no_code[:500],
                signal_type=SignalType.INFO,
                confidence=0.5,
                source_line=line_num,
                source_format='ai_chat',
                tags=['context']
            ))
        
        return blocks
    
    def _process_email(self, content: str) -> List[DistilledBlock]:
        """Process email content."""
        blocks = []
        
        # Extract headers
        header_match = re.search(
            r'Subject:\s*(.+?)(?:\n|$)',
            content,
            re.IGNORECASE
        )
        if header_match:
            blocks.append(DistilledBlock(
                content=header_match.group(1).strip(),
                signal_type=SignalType.INFO,
                confidence=0.9,
                source_line=1,
                source_format='email',
                tags=['subject']
            ))
        
        # Find action items
        for pattern in self.EMAIL_PATTERNS['action']:
            matches = re.findall(pattern, content, re.IGNORECASE)
            for match in matches:
                blocks.append(DistilledBlock(
                    content=match.strip(),
                    signal_type=SignalType.ACTION,
                    confidence=0.7,
                    source_line=0,
                    source_format='email',
                    tags=['action', 'email']
                ))
        
        return blocks
    
    def _process_markdown(self, content: str) -> List[DistilledBlock]:
        """Process markdown content."""
        from services.kb_parser import get_parser
        
        parser = get_parser()
        parsed = parser.parse_content(content, 'markdown')
        
        blocks = []
        for pb in parsed:
            signal_type = SignalType.INFO
            if pb.block_type == 'heading':
                signal_type = SignalType.INFO
            elif pb.block_type == 'code':
                signal_type = SignalType.CODE
            elif pb.block_type == 'task':
                signal_type = SignalType.ACTION
            
            blocks.append(DistilledBlock(
                content=pb.content,
                signal_type=signal_type,
                confidence=0.8,
                source_line=pb.line_start,
                source_format='markdown',
                tags=pb.tags
            ))
        
        return blocks
    
    def _process_code(self, content: str) -> List[DistilledBlock]:
        """Process code content."""
        from services.kb_parser import get_parser
        
        parser = get_parser()
        parsed = parser.parse_content(content, 'python')
        
        blocks = []
        for pb in parsed:
            blocks.append(DistilledBlock(
                content=pb.content,
                signal_type=SignalType.CODE,
                confidence=0.9,
                source_line=pb.line_start,
                source_format='code',
                tags=pb.tags,
                metadata=pb.metadata
            ))
        
        return blocks
    
    def _process_generic(self, content: str) -> List[DistilledBlock]:
        """Process generic text content."""
        blocks = []
        paragraphs = content.split('\n\n')
        
        line = 1
        for para in paragraphs:
            para = para.strip()
            if not para:
                continue
            
            if self._is_noise(para):
                line += para.count('\n') + 2
                continue
            
            blocks.append(DistilledBlock(
                content=para[:500],
                signal_type=SignalType.INFO,
                confidence=0.5,
                source_line=line,
                source_format='generic',
                tags=[]
            ))
            line += para.count('\n') + 2
        
        return blocks
    
    # ==========================================
    # DATA ANALYSIS METHODS (Deterministic, No LLM)
    # ==========================================
    
    def _process_csv(self, content: str) -> List[DistilledBlock]:
        """Process CSV content with automatic statistics."""
        import csv
        import io
        
        blocks = []
        
        try:
            # Parse CSV
            reader = csv.DictReader(io.StringIO(content))
            rows = list(reader)
            
            if not rows:
                return blocks
            
            # Extract column names
            columns = list(rows[0].keys())
            
            blocks.append(DistilledBlock(
                content=f"Columns: {', '.join(columns)}",
                signal_type=SignalType.INFO,
                confidence=1.0,
                source_line=1,
                source_format='csv',
                tags=['schema'],
                metadata={'columns': columns, 'row_count': len(rows)}
            ))
            
            # Calculate statistics for numeric columns
            numeric_stats = self._calculate_numeric_stats(rows, columns)
            for col, stats in numeric_stats.items():
                blocks.append(DistilledBlock(
                    content=f"{col}: mean={stats['mean']:.2f}, min={stats['min']:.2f}, max={stats['max']:.2f}, std={stats['std']:.2f}",
                    signal_type=SignalType.METRIC,
                    confidence=1.0,
                    source_line=1,
                    source_format='csv',
                    tags=['statistic', col],
                    metadata=stats
                ))
                
                # Detect outliers (values > 2 std from mean)
                for i, row in enumerate(rows, 2):
                    try:
                        val = float(row.get(col, 0))
                        if abs(val - stats['mean']) > 2 * stats['std']:
                            blocks.append(DistilledBlock(
                                content=f"Outlier in {col}: {val} (row {i})",
                                signal_type=SignalType.OUTLIER,
                                confidence=0.8,
                                source_line=i,
                                source_format='csv',
                                tags=['outlier', col],
                                metadata={'column': col, 'value': val, 'row': i}
                            ))
                    except (ValueError, TypeError):
                        pass
            
            # Detect trends in time-series columns
            trends = self._detect_trends(rows, columns)
            for trend in trends:
                blocks.append(DistilledBlock(
                    content=trend['description'],
                    signal_type=SignalType.TREND,
                    confidence=trend['confidence'],
                    source_line=1,
                    source_format='csv',
                    tags=['trend', trend['column']],
                    metadata=trend
                ))
            
        except Exception as e:
            blocks.append(DistilledBlock(
                content=f"CSV parse error: {str(e)}",
                signal_type=SignalType.ERROR,
                confidence=1.0,
                source_line=1,
                source_format='csv',
                tags=['error']
            ))
        
        return blocks
    
    def _process_json(self, content: str) -> List[DistilledBlock]:
        """Process JSON content (API responses, configs, data)."""
        import json as json_lib
        
        blocks = []
        
        try:
            data = json_lib.loads(content)
            
            # Handle array of objects (common data format)
            if isinstance(data, list) and len(data) > 0 and isinstance(data[0], dict):
                # Treat as table-like data
                columns = list(data[0].keys())
                
                blocks.append(DistilledBlock(
                    content=f"JSON Array: {len(data)} objects, keys: {', '.join(columns[:10])}",
                    signal_type=SignalType.INFO,
                    confidence=1.0,
                    source_line=1,
                    source_format='json',
                    tags=['array', 'schema'],
                    metadata={'count': len(data), 'keys': columns}
                ))
                
                # Calculate stats
                numeric_stats = self._calculate_numeric_stats(data, columns)
                for col, stats in numeric_stats.items():
                    blocks.append(DistilledBlock(
                        content=f"{col}: mean={stats['mean']:.2f}, range=[{stats['min']:.2f}, {stats['max']:.2f}]",
                        signal_type=SignalType.METRIC,
                        confidence=1.0,
                        source_line=1,
                        source_format='json',
                        tags=['statistic', col],
                        metadata=stats
                    ))
            
            # Handle nested object (config-like)
            elif isinstance(data, dict):
                self._extract_json_signals(data, blocks, '')
            
        except Exception as e:
            blocks.append(DistilledBlock(
                content=f"JSON parse error: {str(e)}",
                signal_type=SignalType.ERROR,
                confidence=1.0,
                source_line=1,
                source_format='json',
                tags=['error']
            ))
        
        return blocks
    
    def _extract_json_signals(self, obj: dict, blocks: List[DistilledBlock], prefix: str):
        """Recursively extract signals from nested JSON."""
        for key, value in obj.items():
            path = f"{prefix}.{key}" if prefix else key
            
            if isinstance(value, dict):
                self._extract_json_signals(value, blocks, path)
            elif isinstance(value, list):
                blocks.append(DistilledBlock(
                    content=f"{path}: array of {len(value)} items",
                    signal_type=SignalType.DATA,
                    confidence=0.9,
                    source_line=1,
                    source_format='json',
                    tags=['array'],
                    metadata={'path': path, 'count': len(value)}
                ))
            elif isinstance(value, (int, float)):
                blocks.append(DistilledBlock(
                    content=f"{path}: {value}",
                    signal_type=SignalType.METRIC,
                    confidence=0.9,
                    source_line=1,
                    source_format='json',
                    tags=['numeric'],
                    metadata={'path': path, 'value': value}
                ))
    
    def _calculate_numeric_stats(self, rows: List[dict], columns: List[str]) -> Dict[str, dict]:
        """Calculate mean, min, max, std for numeric columns."""
        import math
        
        stats = {}
        
        for col in columns:
            values = []
            for row in rows:
                try:
                    val = float(row.get(col, ''))
                    if not math.isnan(val) and not math.isinf(val):
                        values.append(val)
                except (ValueError, TypeError):
                    continue
            
            if len(values) >= 3:  # Need at least 3 values for meaningful stats
                mean = sum(values) / len(values)
                variance = sum((x - mean) ** 2 for x in values) / len(values)
                std = math.sqrt(variance)
                
                stats[col] = {
                    'column': col,
                    'mean': mean,
                    'min': min(values),
                    'max': max(values),
                    'std': std,
                    'count': len(values)
                }
        
        return stats
    
    def _detect_trends(self, rows: List[dict], columns: List[str]) -> List[dict]:
        """Detect simple trends in time-series data."""
        trends = []
        
        for col in columns:
            values = []
            for row in rows:
                try:
                    val = float(row.get(col, ''))
                    values.append(val)
                except (ValueError, TypeError):
                    continue
            
            if len(values) >= 5:
                # Simple trend detection: compare first half vs second half
                mid = len(values) // 2
                first_half_avg = sum(values[:mid]) / mid
                second_half_avg = sum(values[mid:]) / (len(values) - mid)
                
                change_pct = ((second_half_avg - first_half_avg) / first_half_avg * 100) if first_half_avg != 0 else 0
                
                if abs(change_pct) > 10:  # Significant change threshold
                    direction = "increasing" if change_pct > 0 else "decreasing"
                    trends.append({
                        'column': col,
                        'direction': direction,
                        'change_percent': change_pct,
                        'description': f"{col}: {direction} trend ({change_pct:+.1f}%)",
                        'confidence': min(abs(change_pct) / 50, 0.95)  # Higher change = higher confidence
                    })
        
        return trends
    
    async def process_spreadsheet(self, file_path: str) -> List[DistilledBlock]:
        """Process Excel/LibreOffice Calc files (requires openpyxl or xlrd)."""
        blocks = []
        
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, data_only=True)
            
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                
                # Extract headers from first row
                headers = []
                for cell in sheet[1]:
                    if cell.value:
                        headers.append(str(cell.value))
                
                if not headers:
                    continue
                
                blocks.append(DistilledBlock(
                    content=f"Sheet '{sheet_name}': {len(headers)} columns, ~{sheet.max_row} rows",
                    signal_type=SignalType.INFO,
                    confidence=1.0,
                    source_line=1,
                    source_format='spreadsheet',
                    tags=['sheet', sheet_name],
                    metadata={'sheet': sheet_name, 'columns': headers, 'rows': sheet.max_row}
                ))
                
                # Convert to dict list for analysis
                rows = []
                for row in sheet.iter_rows(min_row=2, max_row=min(sheet.max_row, 1000)):
                    row_dict = {}
                    for i, cell in enumerate(row):
                        if i < len(headers):
                            row_dict[headers[i]] = cell.value
                    rows.append(row_dict)
                
                # Calculate statistics
                numeric_stats = self._calculate_numeric_stats(rows, headers)
                for col, stats in numeric_stats.items():
                    blocks.append(DistilledBlock(
                        content=f"[{sheet_name}] {col}: mean={stats['mean']:.2f}, range=[{stats['min']:.2f}, {stats['max']:.2f}]",
                        signal_type=SignalType.METRIC,
                        confidence=1.0,
                        source_line=1,
                        source_format='spreadsheet',
                        tags=['statistic', sheet_name, col],
                        metadata={**stats, 'sheet': sheet_name}
                    ))
                
                # Detect trends
                trends = self._detect_trends(rows, headers)
                for trend in trends:
                    blocks.append(DistilledBlock(
                        content=f"[{sheet_name}] {trend['description']}",
                        signal_type=SignalType.TREND,
                        confidence=trend['confidence'],
                        source_line=1,
                        source_format='spreadsheet',
                        tags=['trend', sheet_name, trend['column']],
                        metadata={**trend, 'sheet': sheet_name}
                    ))
            
            wb.close()
            
        except ImportError:
            blocks.append(DistilledBlock(
                content="openpyxl not installed. Run: pip install openpyxl",
                signal_type=SignalType.ERROR,
                confidence=1.0,
                source_line=1,
                source_format='spreadsheet',
                tags=['dependency']
            ))
        except Exception as e:
            blocks.append(DistilledBlock(
                content=f"Spreadsheet error: {str(e)}",
                signal_type=SignalType.ERROR,
                confidence=1.0,
                source_line=1,
                source_format='spreadsheet',
                tags=['error']
            ))


# Singleton instance
_distiller: Optional[Distiller] = None

def get_distiller() -> Distiller:
    global _distiller
    if _distiller is None:
        _distiller = Distiller()
    return _distiller
